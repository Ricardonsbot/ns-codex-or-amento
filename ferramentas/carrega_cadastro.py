"""Carrega o cadastro REAL da NSTECH dos .xlsx para os JSON de Referencias/.

Ferramenta de bastidor, como gera_inputs.py: roda na mão, e o que entra no
repositório é só o JSON. O site nunca lê .xlsx.

Fontes (as duas ficam em Referencias/, versionadas junto):
  tbl_KMM_Contas.xlsx        -> dContas, 635 contas do plano de contas
  tbl_KMM_Organizacional.xlsx -> Organizacional, 3.284 linhas de estrutura

Escreve:
  contas.json         plano de contas orçável
  centros-custo.json  centros de custo com código e nome
  organizacional.json empresas (preserva a hierarquia já curada, acrescenta o resto)

Três decisões embutidas aqui, porque são de tradução e não de dado:

1. CONTA DE BALANÇO FICA DE FORA. Das 635, 231 são Ativo, Passivo ou tributo
   diferido. Não se orça saldo de balanço — orça-se P&L e Capex. Sobram 404.

2. FPA_Pacote VIRA "Linha P&L", NÃO VIRA PACOTE. É a armadilha desta planilha:
   a coluna se chama pacote mas guarda a linha do P&L (Personnel Costs, Third
   Party Services & Mkt...). Pacote neste produto é o MOTIVO do gasto — Operação
   Base, Novos Contratos, Eficiência — e isso não existe em lugar nenhum do
   .xlsx. pacotes.json continua como está, escrito à mão.

3. A DESCRIÇÃO VEM COMO ESTÁ NO ERP, em caixa alta. Quem lança procura a conta
   pelo texto que vê no sistema de origem; "traduzir" para caixa e baixa faria a
   busca falhar justamente para quem conhece o plano de contas.

O que esta planilha NÃO resolve: o de/para Torre -> BU. As empresas novas entram
com torre "-" e ficam esperando as ~12 linhas que só a área consegue escrever.
"""
import json
import os
import re

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF = os.path.join(RAIZ, "Referencias")

# FPA_Pacote destes grupos é conta patrimonial: não entra no orçamento.
BALANCO = {
    "Ativo", "Passivo", "Caixa e equivalentes de caixa", "Tributos a pagar",
    "Tributos a recuperar", "Imobilizado", "Adiantamentos", "Despesas antecipadas",
    "Direito de Uso", "Outros passivos", "Bônus a pagar", "Fornecedores",
    "Passivo de arrendamento", "Salarios e encargos a pagar", "Intercompany",
    "Imposto de Renda e Contribuição Social a pagar",
    "Contas a pagar em combinação de negócios",
}
RECEITA = {"Gross Revenue", "(-) Deductions"}
VAZIOS = {"", "NA", "NDA", "0", "SemClass", "SemDep"}


def limpo(valor):
    """Texto da planilha sem espaço duplo e sem espaço fino (\xa0) do Excel."""
    return re.sub(r"\s+", " ", str(valor or "").replace("\xa0", " ")).strip()


def linhas_da_aba(caminho, aba):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba]
    it = ws.iter_rows(values_only=True)
    cabecalho = [limpo(c) for c in next(it)]
    for linha in it:
        registro = {cabecalho[i]: limpo(v) for i, v in enumerate(linha) if i < len(cabecalho)}
        if registro.get(cabecalho[0]):
            yield registro
    wb.close()


def gravar(nome, conteudo):
    with open(os.path.join(REF, nome), "w", encoding="utf-8") as fh:
        json.dump(conteudo, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


# ---------------------------------------------------------------- contas

def carregar_contas():
    contas, descartadas = [], 0

    for r in linhas_da_aba(os.path.join(REF, "tbl_KMM_Contas.xlsx"), "dContas"):
        pacote = r.get("FPA_Pacote", "")
        if pacote in BALANCO or not pacote:
            descartadas += 1
            continue

        if pacote in RECEITA:
            linha_pl = f"Receita > {pacote}"
        elif pacote == "Capex":
            # Capex se organiza por tipo de ativo, não por linha de resultado
            linha_pl = f"Capex > {r.get('FPA_Subpacote') or pacote}"
        else:
            linha_pl = f"Despesas > {pacote}"

        contas.append({
            "conta": r["ContaContabil"],
            "nome": r.get("ContaContabil_Descricao", ""),
            "linhaPL": linha_pl,
            "categoria": r.get("FPA_Subpacote") or pacote,
        })

    contas.sort(key=lambda c: c["conta"])
    gravar("contas.json", contas)
    print(f"contas.json          {len(contas):5d} contas orçáveis ({descartadas} de balanço fora)")
    return contas


# --------------------------------------------------- centros de custo

def carregar_centros(org):
    vistos, centros = set(), []

    for r in org:
        codigo = r.get("Nstech_CentroCusto_Codigo", "")
        nome = r.get("Nstech_CentroCusto_Nome", "")
        if codigo in VAZIOS or nome in VAZIOS or codigo in vistos:
            continue
        vistos.add(codigo)
        centros.append({
            "codigo": codigo,
            "nome": nome,
            "empresa": r.get("Organizacional_Empresa", ""),
            "categoriaPL": r.get("PnL_Category", ""),
        })

    centros.sort(key=lambda c: (c["empresa"], c["codigo"]))
    empresas = sorted({c["empresa"] for c in centros})
    gravar("centros-custo.json", {
        "observacao": (
            "Centros de custo reais, de tbl_KMM_Organizacional.xlsx (colunas "
            "Nstech_CentroCusto_Codigo e _Nome). A planilha só tem centro de custo "
            "cadastrado para " + " e ".join(empresas) + " — as demais empresas ainda "
            "lançam sem centro de custo. Gerado por ferramentas/carrega_cadastro.py."
        ),
        "empresasComCadastro": empresas,
        "centros": centros,
    })
    print(f"centros-custo.json   {len(centros):5d} centros ({', '.join(empresas)})")
    return centros


# ------------------------------------------------------- organizacional

# Onde cada empresa mora. Procedência de cada linha, porque nem tudo aqui é
# leitura de planilha:
#   [D] deparaempresas.xlsx nomeia a empresa na coluna Produto_BaseRecebimento
#   [X] tbl_KMM_Organizacional.xlsx aponta o produto da empresa
#   [O] curadoria que já estava em organizacional.json, reencaixada na torre nova
#   [?] inferência por grupo econômico — vale conferir com a área
#
# As três torres de Corporate (G&A, S&M, R&D) não vêm do de/para: ele só tem
# "Consolidado Nstech". Ficam como estavam, porque entregas e aprovações já
# apontam para elas.
EMPRESA_TORRE = [
    # ---- PSL
    ("Bsoft",                      "Torre ICP Pequeno e Micro", "-"),              # [O]
    ("Datamex",                    "Torre ICP Pequeno e Micro", "-"),              # [O]
    ("Hivecloud",                  "Torre ICP Pequeno e Micro", "-"),              # [O]
    ("Otimizy",                    "Torre ICP Pequeno e Micro", "-"),              # [O]
    ("KMM",                        "Torre ICP Médio e Grande",  "-"),              # [O]
    ("KMM - Gridnet",              "Torre ICP Médio e Grande",  "-"),              # [?]
    ("Praxio - Avacorp",           "Torre ICP Médio e Grande",  "-"),              # [O]
    ("Praxio - Globus",            "Torre ICP Médio e Grande",  "-"),              # [O]
    ("Praxio - Compartilhado",     "Torre ICP Médio e Grande",  "Compartilhado"),  # [?]
    ("Praxio - Corporate S&M",     "Torre ICP Médio e Grande",  "Compartilhado"),  # [?]
    ("ATS - Logistica",            "Torre ICP Médio e Grande",  "-"),              # [O]
    ("ATS - Jornada",              "Torre ICP Médio e Grande",  "Jornada"),        # [D]
    ("ATS - Compartilhado",        "Torre ICP Médio e Grande",  "Compartilhado"),  # [?]
    ("ATS - Corporate",            "Torre ICP Médio e Grande",  "Compartilhado"),  # [?]
    ("99Kote",                     "Torre ICP Médio e Grande",  "nskote"),         # [D] "nskote"
    ("Signa",                      "TMS Cabotagem",             "-"),              # [X] produto TMS Cabotagem
    ("Gasola",                     "Fintech",                   "-"),              # [D]
    ("e-frete",                    "Fintech",                   "e-frete nsflow"), # [D]
    ("Digitalcomm",                "Fintech",                   "nspedágio"),      # [D]
    ("Qualp",                      "Fintech",                   "nspedágio"),      # [D]
    ("Praxio - Mobilidade",        "Torre Mobilidade",          "-"),              # [D]
    ("Praxio - Mobilidade Globus", "Torre Mobilidade",          "-"),              # [X]
    ("Buonny",                     "Buonny",                    "-"),              # [D]
    ("Buonny - BPS",               "Buonny",                    "BPS"),            # [D] "BPS"
    ("Atua Redes",                 "Atua Redes",                "-"),              # [D] "Redes"
    ("Atua Sistemas",              "Atua Redes",                "-"),              # [?]
    ("Atua - Compartilhado",       "Atua Redes",                "Compartilhado"),  # [?]
    # ---- Embarcador
    ("Opentech",                   "Torre VGR",                 "Núcleo VGR"),     # [O]
    ("BRK",                        "Torre VGR",                 "Núcleo VGR"),     # [O]
    ("LogRisk",                    "Torre VGR",                 "Núcleo VGR"),     # [O]
    ("Onisys",                     "Torre VGR",                 "Núcleo VGR"),     # [O]
    ("Trafegus",                   "Torre VGR",                 "Núcleo VGR"),     # [O]
    ("GBM Consultoria",            "YMS/WMS",                   "-"),              # [O]
    ("GBM - Consultoria",          "YMS/WMS",                   "-"),              # [?] grafia
    ("Gbm Consultoria",            "YMS/WMS",                   "-"),              # [?] grafia
    ("GBM Logística",              "YMS/WMS",                   "-"),              # [O]
    ("LogOne",                     "YMS/WMS",                   "-"),              # [O]
    ("Multisoftware",              "Torre SW Embarcador",       "-"),              # [D]
    ("Fusion",                     "Torre SW Embarcador",       "Last Mile"),      # [D]
    ("RoutEasy",                   "Torre SW Embarcador",       "Last Mile"),      # [D]
    ("Comprovei",                  "Torre SW Embarcador",       "Last Mile"),      # [D]
    ("Frete Rapido",               "Torre SW Embarcador",       "-"),              # [D]
    ("Runtec",                     "Torre SW Embarcador",       "-"),              # [D]
    ("Trizy",                      "Torre SW Embarcador",       "Trizy App"),      # [D]
    ("Mundo Logistica",            "Mídia",                     "-"),              # [D]
    ("IM",                         "Insurance Market",          "-"),              # [D]
    # ---- Corporate (fora do de/para)
    ("Nstech Corporate Holding",   "G&A",                       "-"),
    ("Nstech Corporate Holding",   "S&M",                       "-"),
    ("Nstech Corporate Holding",   "R&D",                       "-"),
    ("Nstech",                     "G&A",                       "-"),              # [X] holding
    ("Nstech GR",                  "G&A",                       "-"),              # [X]
    ("Nstech S.A.",                "G&A",                       "-"),              # [X]
]

TORRES_CORPORATE = {"G&A", "S&M", "R&D"}

# Torre nova -> torre de onde ela saiu, para herdar o líder já cadastrado.
LIDER_HERDADO = {
    "Torre ICP Pequeno e Micro": "Torre TMS",
    "Torre ICP Médio e Grande":  "Torre TMS",
    "TMS Cabotagem":             "Torre TMS",
    "Fintech":                   "Torre Fintech",
    "Buonny":                    "Torre Buonny",
    "YMS/WMS":                   "Torre SW Embarcador",
    "Atua Redes":                "-",            # Atua ficava no guarda-chuva Corporate
}


def ler_depara():
    """deparaempresas.xlsx: BU / Torre / Produto_BaseRecebimento.

    A planilha tem linhas em branco separando as BUs e linhas de subtotal
    ("Consolidado PSL"), que não são estrutura — são totalizador de relatório.
    """
    wb = openpyxl.load_workbook(os.path.join(REF, "deparaempresas.xlsx"),
                                read_only=True, data_only=True)
    ws = wb["Sheet1"]

    bu_da_torre, produtos_da_torre, ordem = {}, {}, []
    for linha in ws.iter_rows(values_only=True):
        valores = [limpo(linha[i]) if i < len(linha) else "" for i in range(3)]
        bu, torre, produto = valores
        if not torre or torre == "Torre" or bu.startswith("Consolidado"):
            continue
        if torre not in bu_da_torre:
            bu_da_torre[torre] = bu
            produtos_da_torre[torre] = []
            ordem.append(torre)
        if produto and produto not in produtos_da_torre[torre]:
            produtos_da_torre[torre].append(produto)
    wb.close()

    porbu = {}
    for t, b in bu_da_torre.items():
        porbu.setdefault(b, []).append(t)
    print("de/para lido:", " | ".join(f"{b}: {len(t)} torres" for b, t in porbu.items()))
    return bu_da_torre, produtos_da_torre, ordem


def carregar_organizacional(org, bu_da_torre):
    """BU -> Torre -> Sub Torre -> Empresa, com a torre vinda do de/para."""
    catalogo = {r["Organizacional_Empresa"] for r in org
                if r.get("Organizacional_Empresa") not in VAZIOS}

    saida, mapeadas = [], set()
    for empresa, torre, sub in EMPRESA_TORRE:
        bu = bu_da_torre.get(torre, "Corporate" if torre in TORRES_CORPORATE else "-")
        saida.append({"bu": bu, "torre": torre, "subTorre": sub, "empresa": empresa})
        mapeadas.add(empresa)

    # empresa que aparece no ERP e ninguém encaixou: entra sem torre, visível
    orfas = sorted(catalogo - mapeadas)
    saida += [{"bu": "-", "torre": "-", "subTorre": "-", "empresa": e} for e in orfas]

    gravar("organizacional.json", saida)
    torres = {t for _, t, _ in EMPRESA_TORRE}
    print(f"organizacional.json  {len(saida):5d} linhas | {len(mapeadas)} empresas em {len(torres)} torres"
          + (f" | {len(orfas)} sem torre: {orfas}" if orfas else " | nenhuma sem torre"))
    return saida


def carregar_produtos(produtos_da_torre, ordem):
    """O produto do de/para vira o catálogo da tela de Receita."""
    with open(os.path.join(REF, "produtos.json"), encoding="utf-8") as fh:
        atual = json.load(fh)

    torres_do_produto = {}
    for torre in ordem:
        for p in produtos_da_torre[torre]:
            torres_do_produto.setdefault(p, []).append(torre)

    produtos = [{"nome": nome, "torres": torres, "subProdutos": []}
                for nome, torres in torres_do_produto.items()]

    gravar("produtos.json", {
        "observacao": (
            "Estrutura de input da Receita: Torre -> Empresa -> Produto -> Sub-produto -> "
            "Tipo de Receita. Torre e Empresa vem de organizacional.json; o catalogo de "
            "produto vem de deparaempresas.xlsx (coluna Produto_BaseRecebimento), a mesma "
            "base de recebimento que o financeiro usa. Sub-produto ficou vazio: a planilha "
            "para no produto. Gerado por ferramentas/carrega_cadastro.py."
        ),
        "produtos": produtos,
    })
    print(f"produtos.json        {len(produtos):5d} produtos em {len(ordem)} torres")
    return produtos


def realinhar_derivados(organizacional):
    """entregas.json e aprovacoes.json guardam bu/torre próprios. Se a torre muda
    aqui e não muda lá, a tela de Entregas mostra uma árvore e o formulário mostra
    outra — que é exatamente o que este projeto não pode ter."""
    onde = {}
    for e in organizacional:
        onde.setdefault(e["empresa"], (e["bu"], e["torre"], e["subTorre"]))

    trocas = 0

    with open(os.path.join(REF, "entregas.json"), encoding="utf-8") as fh:
        entregas = json.load(fh)
    for e in entregas["entregas"]:
        bu, torre, sub = onde.get(e["empresa"], (e["bu"], e["torre"], e["subTorre"]))
        trocas += (e["bu"], e["torre"]) != (bu, torre)
        e["bu"], e["torre"], e["subTorre"] = bu, torre, sub
    gravar("entregas.json", entregas)

    with open(os.path.join(REF, "aprovacoes.json"), encoding="utf-8") as fh:
        aprov = json.load(fh)
    for s in aprov["submissoes"]:
        bu, torre, _ = onde.get(s["empresa"], (s["bu"], s["torre"], "-"))
        trocas += (s["bu"], s["torre"]) != (bu, torre)
        s["bu"], s["torre"] = bu, torre

    # O líder é cadastrado por torre. Com o de/para, a Torre TMS virou três
    # torres e a SW Embarcador soltou a YMS/WMS — quem liderava a torre de
    # origem segue respondendo pelas que saíram dela, até a área dizer outra
    # coisa. Sem isto a tela de Aprovações fica sem aprovador nessas torres.
    antigos, lideres, sem_lider = aprov["lideres"], {}, []
    for torre in sorted({s["torre"] for s in aprov["submissoes"]}):
        origem = torre if torre in antigos else LIDER_HERDADO.get(torre)
        if origem and origem in antigos:
            lider = dict(antigos[origem])
            preposicao = "da" if torre.startswith("Torre") else "de"
            lider["cargo"] = f"Líder {preposicao} {torre}"
            lideres[torre] = lider
        else:
            sem_lider.append(torre)
    aprov["lideres"] = lideres

    # O aceite final guarda quem assinou e com que cargo, congelado no momento
    # do aceite. Com a torre renomeada o cargo virava fóssil ("Líder da Torre
    # TMS" numa entrega que agora é da Torre ICP Médio e Grande) e aparecia
    # assim na tela de Aprovações. Dado sintético de protótipo: reescreve.
    for s in aprov["submissoes"]:
        lider = lideres.get(s["torre"])
        if not lider:
            continue
        if s.get("liderResponsavel"):
            s["liderResponsavel"] = dict(lider)
        if s.get("aceiteFinal"):
            s["aceiteFinal"]["por"], s["aceiteFinal"]["cargo"] = lider["nome"], lider["cargo"]

    gravar("aprovacoes.json", aprov)

    print(f"entregas/aprovacoes  {trocas:5d} linhas com BU/Torre reescritos"
          + (f"\n  >> torres sem líder cadastrado: {sem_lider}" if sem_lider else ""))
    return sem_lider


if __name__ == "__main__":
    org = list(linhas_da_aba(os.path.join(REF, "tbl_KMM_Organizacional.xlsx"), "Organizacional"))
    print(f"lidas {len(org)} linhas organizacionais")
    bu_da_torre, produtos_da_torre, ordem = ler_depara()
    print()

    carregar_contas()
    carregar_centros(org)
    organizacional = carregar_organizacional(org, bu_da_torre)
    carregar_produtos(produtos_da_torre, ordem)
    realinhar_derivados(organizacional)

    print("\nagora rode: python ferramentas/gera_inputs.py")
