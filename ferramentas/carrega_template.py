"""Extrai do Template Budget as dimensões que faltavam no input da plataforma.

Ferramenta de bastidor, como as outras: roda na mão, e o que entra no
repositório é só o JSON.

Fonte: Referencias/template-budget-psl.xlsx — o arquivo que as áreas preenchem
hoje. As abas "Base Receita" e "Base Gastos" são o desenho ideal do input, e
este script traz os conceitos delas para cá.

O QUE O TEMPLATE ENSINOU

1. Receita se lança por CONTRATO, não por produto. A linha tem CNPJ e razão
   social do cliente; produto é atributo dela. São 3.798 clientes.

2. "Tipo Receita" e "Categoria" são DOIS EIXOS, não um. Movimento (é base? é
   venda nova? é churn?) e natureza (SaaS, On Premise, Implantação). A
   plataforma só tinha um campo, com o nome de um e os valores do outro.

3. Reajuste tem aniversário. Cada contrato traz o índice e o mês em que ele
   incide — é isso que permite diluir o efeito ao longo do ano em vez de
   aplicar tudo em janeiro.

4. O "Pacote" do template é a linha do P&L, não o motivo do gasto. Mesma
   armadilha do FPA_Pacote. Aqui ele NÃO vira pacote: a linha do P&L já sai da
   conta, e o pacote-motivo continua sendo campo próprio da plataforma.

Escreve:
  clientes.json          carteira com CNPJ, razão social e os atributos do cliente
  dimensoes-receita.json os eixos de classificação da receita
  indices-reajuste.json  índice acumulado por mês de aniversário
  fornecedores.json      fornecedores e projetos vistos na Base Gastos
"""
import json
import os
import re
from collections import Counter, OrderedDict

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF = os.path.join(RAIZ, "Referencias")
TEMPLATE = os.path.join(REF, "template-budget-psl.xlsx")

# A coluna Empresa do template não bate com o cadastro: vem com grafia livre e,
# em algumas linhas, com Torre ou BU no lugar da empresa. O de/para é aqui
# porque é defeito de digitação na origem, não conceito.
EMPRESA_DEPARA = {
    "kmm": "KMM", "bsoft": "Bsoft", "99kote": "99Kote", "hive": "Hivecloud",
    "atua": "Atua Sistemas", "ats jornada": "ATS - Jornada",
    "ats logistica": "ATS - Logistica", "e-frete nsflow": "e-frete",
    "e-frete pedagio": "e-frete", "e-frete meio pagamento": "e-frete",
    "praxio - mobilidade": "Praxio - Mobilidade", "torre vgr buonny": "Buonny",
}
# Estas não são empresa — são torre ou BU na coluna errada. Ficam registradas
# para o import avisar em vez de aceitar calado.
NAO_E_EMPRESA = {"bu psl - cs", "bu psl - comercial", "bu psl - g&a",
                 "torre mg", "marketplace"}

VAZIOS = {"", "0", "nan", "none", "-"}


def limpo(v):
    return re.sub(r"\s+", " ", str(v or "").replace("\xa0", " ")).strip()


def normaliza_empresa(bruto):
    """Devolve (empresa_do_cadastro, aviso). Aviso preenchido = precisa de olho."""
    nome = limpo(bruto)
    chave = nome.lower()
    if chave in NAO_E_EMPRESA:
        return nome, "nao e empresa"
    return EMPRESA_DEPARA.get(chave, nome), None


def gravar(nome, conteudo):
    with open(os.path.join(REF, nome), "w", encoding="utf-8") as fh:
        json.dump(conteudo, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


def valores(linhas, i):
    """Valores distintos de uma coluna, na ordem de frequência."""
    c = Counter(limpo(l[i]) for l in linhas if limpo(l[i]).lower() not in VAZIOS)
    return [v for v, _ in c.most_common()]


def numero(v):
    return v if isinstance(v, (int, float)) else 0.0


# --------------------------------------------------------------- receita

def carregar_receita(wb):
    linhas = [l for l in wb["Base Receita"].iter_rows(min_row=5, values_only=True) if l[1]]

    # O cliente é a linha do template. Um CNPJ aparece em vários contratos, e
    # os atributos (persona, setor, classe) viajam com ele — fica um registro
    # por CNPJ, com o que aparece mais vezes.
    porCnpj = OrderedDict()
    for l in linhas:
        cnpj = limpo(l[3])
        if not cnpj or cnpj.lower() in VAZIOS:
            continue
        d = porCnpj.setdefault(cnpj, {"contratos": 0, "receita": 0.0,
                                      "campos": {k: Counter() for k in
                                                 ("razao", "empresa", "persona", "setor",
                                                  "segmento", "classe", "pmr")}})
        d["contratos"] += 1
        d["receita"] += numero(l[27])
        emp, _ = normaliza_empresa(l[1])
        for chave, valor in (("razao", l[4]), ("empresa", emp), ("persona", l[8]),
                             ("setor", l[9]), ("segmento", l[10]), ("classe", l[11]),
                             ("pmr", l[5])):
            v = limpo(valor)
            if v and v.lower() not in VAZIOS:
                d["campos"][chave][v] += 1

    clientes = []
    for cnpj, d in porCnpj.items():
        mais = lambda k: (d["campos"][k].most_common(1) or [("", 0)])[0][0]
        clientes.append({
            "cnpj": cnpj, "razaoSocial": mais("razao"), "empresa": mais("empresa"),
            "persona": mais("persona"), "setor": mais("setor"),
            "segmento": mais("segmento"), "classe": mais("classe"),
            "pmr": mais("pmr"), "contratos": d["contratos"],
            "receitaAno": round(d["receita"], 2),
        })
    clientes.sort(key=lambda c: -c["receitaAno"])
    gravar("clientes.json", {
        "observacao": ("Carteira de clientes da Base Receita do template. A receita se "
                       "lanca por contrato: o CNPJ e a linha, e produto e atributo dela. "
                       "Persona, setor, segmento e classe viajam com o cliente. Gerado por "
                       "ferramentas/carrega_template.py."),
        "clientes": clientes,
    })

    dim = {
        "observacao": ("Os dois eixos de classificacao da receita, que o template mantem "
                       "separados e a plataforma juntava num campo so. MOVIMENTO responde "
                       "'de onde vem o numero' (base, venda nova, churn); NATUREZA responde "
                       "'o que foi vendido' (SaaS, On Premise, implantacao). "
                       "Gerado por ferramentas/carrega_template.py."),
        "movimento": valores(linhas, 2),
        "natureza": valores(linhas, 6),
        "personas": valores(linhas, 8),
        "setores": valores(linhas, 9),
        "segmentos": valores(linhas, 10),
        "classesCliente": valores(linhas, 11),
        "termometro": valores(linhas, 12),
    }
    gravar("dimensoes-receita.json", dim)

    print(f"clientes.json          {len(clientes):5d} clientes ({len(linhas)} contratos)")
    print(f"dimensoes-receita.json       movimento={len(dim['movimento'])} "
          f"natureza={len(dim['natureza'])} personas={len(dim['personas'])} "
          f"termometro={len(dim['termometro'])}")
    return linhas


def carregar_indices(wb, linhas_receita):
    """Índice acumulado 12m por mês de aniversário — o que permite diluir o
    reajuste em vez de jogar tudo em janeiro."""
    ws = wb["Indices Reajuste"]
    tabela = list(ws.iter_rows(min_row=3, max_row=15, values_only=True))
    cab = [limpo(c) for c in tabela[0]]
    nomes = [c for c in cab[2:8] if c]

    meses = []
    for l in tabela[1:]:
        if not l[1]:
            continue
        mes = str(l[1])[:7]
        meses.append({"mes": mes, **{nomes[i]: round(numero(l[2 + i]), 6)
                                     for i in range(len(nomes))}})

    usados = Counter(limpo(l[29]) for l in linhas_receita if limpo(l[29]))
    aniversarios = Counter(str(l[30])[:7] for l in linhas_receita if l[30])

    gravar("indices-reajuste.json", {
        "observacao": ("Indice acumulado 12 meses por mes de aniversario, da aba 'Indices "
                       "Reajuste'. Cada contrato reajusta no seu mes, entao o efeito entra "
                       "diluido ao longo do ano. 'Livre' e contrato sem indexacao. "
                       "Gerado por ferramentas/carrega_template.py."),
        "indices": nomes,
        "usoNosContratos": dict(usados),
        "mesesDeAniversario": dict(sorted(aniversarios.items())),
        "meses": meses,
    })
    print(f"indices-reajuste.json  {len(meses):5d} meses x {len(nomes)} indices "
          f"| uso: {dict(usados)}")


# ---------------------------------------------------------------- gastos

def carregar_gastos(wb):
    linhas = [l for l in wb["Base Gastos"].iter_rows(min_row=4, values_only=True) if l[0]]
    extra = [l for l in wb["Base Gastos - Extra"].iter_rows(min_row=4, values_only=True) if l[0]]
    todas = linhas + extra

    fornecedores = valores(todas, 6)
    projetos = valores(todas, 7)
    areas = sorted({limpo(l[26]).replace("Cogs", "CoGS") for l in todas
                    if limpo(l[26]).lower() not in VAZIOS})
    responsaveis = valores(todas, 27)

    avisos = []
    for l in todas:
        _, aviso = normaliza_empresa(l[0])
        if aviso:
            avisos.append(limpo(l[0]))

    gravar("fornecedores.json", {
        "observacao": ("Fornecedores, projetos, areas e responsaveis vistos na Base Gastos "
                       "do template — as colunas que a grade de Despesa nao tinha. "
                       "'Area' e a linha do P&L gerencial (G&A, CoGS, S&M, R&D) e vem com "
                       "grafia inconsistente na origem ('Cogs' x 'CoGS'), normalizada aqui. "
                       "Gerado por ferramentas/carrega_template.py."),
        "fornecedores": fornecedores,
        "projetos": projetos,
        "areas": areas,
        "responsaveis": responsaveis,
    })
    print(f"fornecedores.json      {len(fornecedores):5d} fornecedores | {len(projetos)} projetos "
          f"| areas={areas} | {len(responsaveis)} responsaveis")
    if avisos:
        print(f"  >> {len(avisos)} linha(s) com Torre/BU na coluna Empresa: {sorted(set(avisos))}")
    return todas


if __name__ == "__main__":
    wb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=True)
    receita = carregar_receita(wb)
    carregar_indices(wb, receita)
    carregar_gastos(wb)
    wb.close()
    print("\nO template nao traz Ativacao na Base Gastos (fica em aba separada, so com "
          "realizado 2025) nem Sub-produto. Esses dois seguem sem fonte.")
