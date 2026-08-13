"""Semeia as grades de Receita, Despesa e Capex com base real do template.

Ferramenta de bastidor: roda na mão e regrava o <tbody> das três grades.

POR QUE

Para o período de testes, a grade não pode chegar com meia dúzia de linhas de
exemplo. Quem vai testar precisa encontrar a conta que procura, o cliente que
conhece e o produto que vende — senão a primeira coisa que faz é concluir que
o cadastro está vazio.

O QUE ENTRA

  Despesa  uma linha por conta contábil distinta da Base Gastos, com o
           fornecedor, o centro de custo, a área e o responsável reais
  Receita  contratos reais, com CNPJ e razão social da carteira, cobrindo os
           seis tipos de receita
  Capex    as linhas da Base CAPEX que têm valor

Valores em R$ mil, como o resto do protótipo. Gasto vem negativo na planilha e
entra positivo aqui, porque a plataforma já sabe que despesa subtrai.
"""
import datetime
import html
import io
import json
import os
import re

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF = os.path.join(RAIZ, "Referencias")
TEMPLATE = os.path.join(REF, "template-budget-psl.xlsx")

LINHAS_DESPESA = 60
LINHAS_RECEITA = 60

EMPRESA_DEPARA = {
    "kmm": "KMM", "bsoft": "Bsoft", "99kote": "99Kote", "hive": "Hivecloud",
    "atua": "Atua Sistemas", "ats jornada": "ATS - Jornada",
    "ats logistica": "ATS - Logistica", "e-frete nsflow": "e-frete",
    "e-frete pedagio": "e-frete", "e-frete meio pagamento": "e-frete",
    "torre vgr buonny": "Buonny",
}


def limpo(v):
    return re.sub(r"\s+", " ", str(v or "").replace("\xa0", " ")).strip()


def empresa(v):
    n = limpo(v)
    return EMPRESA_DEPARA.get(n.lower(), n)


def esc(v):
    return html.escape(limpo(v), quote=True)


def mil(v):
    n = v if isinstance(v, (int, float)) else 0
    return round(abs(n) / 1000, 1)


def mes_iso(v):
    try:
        n = float(v)
    except (TypeError, ValueError):
        return limpo(v)[:7]
    if n < 20000:
        return limpo(v)[:7]
    d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)
    return f"{d.year}-{d.month:02d}"


def campo(valor, classe="", lista="", extra=""):
    c = f' class="{classe}"' if classe else ""
    l = f' list="{lista}"' if lista else ""
    return f'<input type="text" value="{esc(valor)}"{c}{l}{extra} />'


class Grade:
    """Reaproveita o markup da primeira linha existente e só troca o conteúdo,
    para nenhum comportamento de grade (duplicar, remover, colar) se perder."""

    def __init__(self, arquivo, tabela):
        self.arquivo = os.path.join(RAIZ, arquivo)
        self.tabela = tabela
        self.html = io.open(self.arquivo, encoding="utf-8").read()
        ini = self.html.index(f'id="{tabela}"')
        modelo = re.search(r"<tbody>\s*(<tr.*?</tr>)", self.html[ini:], re.S).group(1)
        tds = re.findall(r"<td.*?</td>", modelo, re.S)
        self.td_mes = [t for t in tds if 'class="month-col"' in t][0]
        self.td_total = tds[-2]
        self.td_acoes = tds[-1]

    def linha(self, celulas, valores, rotulo):
        partes = [f"<td>{c}</td>" for c in celulas]
        for v in valores[:12]:
            partes.append(re.sub(r'value="[^"]*"', f'value="{v}"', self.td_mes, count=1))
        partes.append(re.sub(r'value="[^"]*"', f'value="{sum(valores):.1f}"', self.td_total, count=1))
        partes.append(re.sub(r'data-action-label="[^"]*"', f'data-action-label="{esc(rotulo)}"', self.td_acoes))
        return "<tr>\n                " + "\n                ".join(partes) + "\n              </tr>"

    def gravar(self, linhas):
        ini = self.html.index(f'id="{self.tabela}"')
        a = self.html.index("<tbody", ini)
        b = self.html.index("</tbody>", ini) + len("</tbody>")
        corpo = "<tbody>\n              " + "\n              ".join(linhas) + "\n            </tbody>"
        io.open(self.arquivo, "w", encoding="utf-8", newline="").write(
            self.html[:a] + corpo + self.html[b:])
        print(f"  {os.path.basename(self.arquivo):26s} {len(linhas):3d} linhas")


def contas_por_codigo():
    with io.open(os.path.join(REF, "contas.json"), encoding="utf-8") as fh:
        return {c["conta"].replace(".", ""): c for c in json.load(fh)}


# ------------------------------------------------------------------ despesa

def semear_gastos(wb, aba, arquivo, tabela, limite, com_projeto, uma_por_conta=True):
    contas = contas_por_codigo()
    linhas = [l for l in wb[aba].iter_rows(min_row=4, values_only=True) if l[0]]
    linhas = [l for l in linhas if any(isinstance(l[9 + i], (int, float)) and l[9 + i] for i in range(12))]

    if uma_por_conta:
        # Despesa: a de maior valor de cada conta, para o teste enxergar o plano
        # inteiro sem a mesma conta repetida dezenas de vezes.
        melhor = {}
        for l in linhas:
            cod = limpo(l[1])
            if cod not in melhor or abs(l[21] or 0) > abs(melhor[cod][21] or 0):
                melhor[cod] = l
        linhas = list(melhor.values())
    # Capex vai inteiro: são 19 linhas em 4 contas, e é o item comprado que
    # distingue uma da outra — deduplicar por conta deixaria a grade com 4.
    escolhidas = sorted(linhas, key=lambda l: -abs(l[21] or 0))[:limite]

    g = Grade(arquivo, tabela)
    saida = []
    for l in escolhidas:
        cod = limpo(l[1])
        c = contas.get(cod, {"conta": cod, "nome": limpo(l[2]), "linhaPL": "", "categoria": ""})
        cc = " — ".join(x for x in (limpo(l[3]), limpo(l[4])) if x)
        celulas = [
            campo(c["conta"], "conta-codigo-input input-readonly", "", ' style="min-width:70px;" readonly'),
            campo(c["nome"], "conta-nome-input", "contas-datalist"),
            campo(c["linhaPL"], "conta-linha-input"),
            campo(c["categoria"], "conta-categoria-input"),
            campo("", "pacote-input", "pacotes-datalist"),
            campo("Não ativa — Opex", "ativacao-input", "ativacao-datalist"),
            '<input type="number" value="0" class="ativacao-pct" min="0" max="100" />',
            campo(cc, "", "centros-datalist"),
            campo(limpo(l[26]).replace("Cogs", "CoGS"), "", "areas-datalist"),
            campo(limpo(l[5]) if limpo(l[5]) not in ("0", "") else "", "", "produtos-datalist"),
            campo(limpo(l[6]), "", "fornecedores-datalist"),
            campo(limpo(l[7]), "", "projetos-datalist"),
            campo(limpo(l[27]), "", "responsaveis-datalist"),
            campo(limpo(l[8])),
        ]
        if com_projeto:      # Capex troca as duas últimas por Projeto e Justificativa
            celulas = celulas[:11] + [campo(limpo(l[7]) or limpo(l[8]), "", "projetos-datalist"),
                                      campo(limpo(l[27]), "", "responsaveis-datalist"),
                                      campo(limpo(l[8]))]
        saida.append(g.linha(celulas, [mil(l[9 + i]) for i in range(12)],
                             f"{empresa(l[0])} · {c['nome'][:24]}"))
    g.gravar(saida)


# ------------------------------------------------------------------ receita

def semear_receita(wb, limite):
    linhas = [l for l in wb["Base Receita"].iter_rows(min_row=5, values_only=True) if l[1]]
    linhas = [l for l in linhas if l[27]]

    # As maiores linhas da base são marcadores sem cliente ("New logo",
    # "CLIENTES D", "Valor Receita Incremental"). Ordenar só por valor encheria
    # a grade delas, e quem for testar precisa justamente ver CNPJ de verdade —
    # então contrato identificado vem primeiro, e o valor é o desempate.
    def identificado(l):
        d = re.sub(r"\D", "", limpo(l[3]))
        return len(d) >= 11 and set(d) != {"0"}

    ordem = lambda l: (0 if identificado(l) else 1, -abs(l[27]))

    # cobre os seis tipos de receita antes de completar pelas maiores
    por_tipo, escolhidas, vistos = {}, [], set()
    for l in linhas:
        por_tipo.setdefault(limpo(l[2]), []).append(l)
    for tipo, lista in por_tipo.items():
        for l in sorted(lista, key=ordem)[:4]:
            chave = (limpo(l[3]), limpo(l[7]), tipo)
            if chave not in vistos:
                escolhidas.append(l); vistos.add(chave)
    for l in sorted(linhas, key=ordem):
        if len(escolhidas) >= limite:
            break
        chave = (limpo(l[3]), limpo(l[7]), limpo(l[2]))
        if chave not in vistos:
            escolhidas.append(l); vistos.add(chave)

    g = Grade("orcamento-receita.html", "tabela-receita")
    saida = []
    for l in escolhidas[:limite]:
        celulas = [
            campo(empresa(l[1]), "", "empresas-datalist"),
            campo(l[2], "", "movimento-datalist"),
            campo(l[3], "cliente-cnpj-input", "cnpjs-datalist", ' style="min-width:120px;"'),
            campo(l[4], "cliente-nome-input", "clientes-datalist"),
            campo(l[5], "input-readonly", "", ' data-cliente-campo="pmr" style="min-width:48px;" readonly'),
            campo(l[6], "", "natureza-datalist"),
            campo(l[7], "", "produtos-datalist"),
            campo(l[8], "", "personas-datalist", ' data-cliente-campo="persona"'),
            campo(l[9], "", "setores-datalist", ' data-cliente-campo="setor"'),
            campo(l[10], "", "segmentos-datalist", ' data-cliente-campo="segmento"'),
            campo(l[11], "", "classes-datalist", ' data-cliente-campo="classe"'),
            campo(l[12], "", "termometro-datalist"),
            campo(l[13], "", "projetos-datalist"),
            campo(l[14]),
            campo(l[29], "indice-input", "indices-datalist"),
            campo(mes_iso(l[30]), "mes-reajuste-input", "mesreajuste-datalist"),
        ]
        valores = [round((l[15 + i] or 0) / 1000, 1) for i in range(12)]
        saida.append(g.linha(celulas, valores, f"{empresa(l[1])} · {limpo(l[4])[:24]}"))
    g.gravar(saida)


if __name__ == "__main__":
    wb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=True)
    semear_receita(wb, LINHAS_RECEITA)
    semear_gastos(wb, "Base Gastos", "orcamento-despesa.html", "tabela-despesa", LINHAS_DESPESA, False)
    semear_gastos(wb, "Base CAPEX", "orcamento-capex.html", "tabela-capex", 30, True, uma_por_conta=False)
    wb.close()
    print("\nDepois disto, rode ferramentas/gera_dados_embutidos.py se algum JSON tiver mudado.")
